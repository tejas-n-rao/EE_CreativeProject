from __future__ import annotations

import json
import os
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT_DIR = Path(__file__).resolve().parents[4]
DEFAULT_FACTOR_JSON_PATH = ROOT_DIR / "data" / "db_emission_factors.json"
DEFAULT_FACTOR_SHEET_PATH = ROOT_DIR / "data" / "emission_factors.xlsx"
FACTOR_SHEET_ENV = "EMISSION_FACTOR_SHEET_PATH"

SHEET_COLUMNS = [
    "activity_name",
    "category",
    "region",
    "unit_activity",
    "default_factor_kgco2e_per_unit",
    "default_source_name",
    "default_source_url",
    "default_source_year",
    "custom_factor_kgco2e_per_unit",
    "custom_source_name",
    "custom_source_url",
    "custom_source_year",
]

ACTIVITY_LABELS: dict[str, str] = {
    "electricity": "Electricity",
    "water_supply_m3": "Water Supply",
    "petrol_car_km": "Petrol Car",
    "diesel_car_km": "Diesel Car",
    "bus_km": "Bus",
    "metro_km": "Metro",
    "rail_km": "Rail",
    "two_wheeler_km": "Two-wheeler",
    "flight_shorthaul": "Short-haul Flight",
    "lpg_kg": "Cooking Gas (LPG)",
    "diet_plant_based_day": "Diet (Plant-based)",
    "diet_mixed_day": "Diet (Mixed)",
    "diet_meat_heavy_day": "Diet (Meat-heavy)",
}


class EmissionFactorSheetError(ValueError):
    pass


@dataclass(frozen=True)
class FactorSheetRow:
    activity_name: str
    category: str
    region: str
    unit_activity: str
    default_factor_kgco2e_per_unit: Decimal
    default_source_name: str
    default_source_url: str
    default_source_year: int | None
    custom_factor_kgco2e_per_unit: Decimal | None
    custom_source_name: str | None
    custom_source_url: str | None
    custom_source_year: int | None

    @property
    def has_custom_override(self) -> bool:
        return self.custom_factor_kgco2e_per_unit is not None


def _sheet_path() -> Path:
    candidate = os.getenv(FACTOR_SHEET_ENV)
    if candidate:
        return Path(candidate).expanduser().resolve()
    return DEFAULT_FACTOR_SHEET_PATH


def _parse_decimal(value: object, *, field_name: str, row_number: int) -> Decimal:
    try:
        result = Decimal(str(value).strip())
    except (InvalidOperation, ValueError, TypeError) as exc:
        raise EmissionFactorSheetError(
            f"Invalid decimal '{field_name}' in emission factor sheet at row {row_number}"
        ) from exc
    return result


def _parse_optional_decimal(value: object) -> Decimal | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError, TypeError):
        return None


def _parse_optional_int(value: object) -> int | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(text)
    except (ValueError, TypeError):
        return None


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _default_activity_name(category: str) -> str:
    return ACTIVITY_LABELS.get(category, category.replace("_", " ").title())


def _bootstrap_factor_sheet(path: Path) -> None:
    if not DEFAULT_FACTOR_JSON_PATH.exists():
        raise EmissionFactorSheetError(
            f"Cannot initialize emission factor sheet; missing {DEFAULT_FACTOR_JSON_PATH}"
        )

    with DEFAULT_FACTOR_JSON_PATH.open("r", encoding="utf-8") as file:
        seed_rows = json.load(file)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "emission_factors"
    worksheet.append(SHEET_COLUMNS)

    for item in seed_rows:
        category = _clean_text(item.get("category"))
        region = _clean_text(item.get("region")) or "WORLD"
        unit_activity = _clean_text(item.get("unit_activity"))
        default_factor = item.get("factor_kgco2e_per_unit")
        default_source_name = _clean_text(item.get("source_name"))
        default_source_url = _clean_text(item.get("source_url"))
        default_source_year = item.get("source_year")

        worksheet.append(
            [
                _default_activity_name(category),
                category,
                region,
                unit_activity,
                default_factor,
                default_source_name,
                default_source_url,
                default_source_year,
                "",
                "",
                "",
                "",
            ]
        )

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def ensure_factor_sheet_exists() -> Path:
    path = _sheet_path()
    if path.exists():
        return path
    _bootstrap_factor_sheet(path)
    return path


def load_factor_sheet_rows() -> list[FactorSheetRow]:
    path = ensure_factor_sheet_exists()
    workbook = load_workbook(BytesIO(path.read_bytes()), data_only=True)
    try:
        worksheet = workbook.active

        header_cells = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        if not header_cells:
            raise EmissionFactorSheetError("Emission factor sheet is empty")

        headers = [_clean_text(value).lower() for value in header_cells[0]]
        key_index = {key: index for index, key in enumerate(headers)}
        required = [
            "activity_name",
            "category",
            "region",
            "unit_activity",
            "default_factor_kgco2e_per_unit",
            "default_source_name",
            "default_source_url",
        ]
        missing = [key for key in required if key not in key_index]
        if missing:
            raise EmissionFactorSheetError(
                f"Emission factor sheet is missing columns: {', '.join(missing)}"
            )

        rows: list[FactorSheetRow] = []
        for row_number, raw in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            category = _clean_text(
                raw[key_index["category"]] if key_index["category"] < len(raw) else ""
            )
            if not category:
                continue

            activity_name = _clean_text(
                raw[key_index["activity_name"]] if key_index["activity_name"] < len(raw) else ""
            ) or _default_activity_name(category)
            region = (
                _clean_text(raw[key_index["region"]] if key_index["region"] < len(raw) else "")
                or "WORLD"
            )
            unit_activity = _clean_text(
                raw[key_index["unit_activity"]] if key_index["unit_activity"] < len(raw) else ""
            )
            default_source_name = _clean_text(
                raw[key_index["default_source_name"]]
                if key_index["default_source_name"] < len(raw)
                else ""
            )
            default_source_url = _clean_text(
                raw[key_index["default_source_url"]]
                if key_index["default_source_url"] < len(raw)
                else ""
            )
            default_factor = _parse_decimal(
                raw[key_index["default_factor_kgco2e_per_unit"]]
                if key_index["default_factor_kgco2e_per_unit"] < len(raw)
                else None,
                field_name="default_factor_kgco2e_per_unit",
                row_number=row_number,
            )

            default_source_year = _parse_optional_int(
                raw[key_index["default_source_year"]]
                if "default_source_year" in key_index and key_index["default_source_year"] < len(raw)
                else None
            )
            custom_factor = _parse_optional_decimal(
                raw[key_index["custom_factor_kgco2e_per_unit"]]
                if "custom_factor_kgco2e_per_unit" in key_index
                and key_index["custom_factor_kgco2e_per_unit"] < len(raw)
                else None
            )
            custom_source_name = _clean_text(
                raw[key_index["custom_source_name"]]
                if "custom_source_name" in key_index and key_index["custom_source_name"] < len(raw)
                else ""
            ) or None
            custom_source_url = _clean_text(
                raw[key_index["custom_source_url"]]
                if "custom_source_url" in key_index and key_index["custom_source_url"] < len(raw)
                else ""
            ) or None
            custom_source_year = _parse_optional_int(
                raw[key_index["custom_source_year"]]
                if "custom_source_year" in key_index and key_index["custom_source_year"] < len(raw)
                else None
            )

            rows.append(
                FactorSheetRow(
                    activity_name=activity_name,
                    category=category,
                    region=region.upper(),
                    unit_activity=unit_activity,
                    default_factor_kgco2e_per_unit=default_factor,
                    default_source_name=default_source_name,
                    default_source_url=default_source_url,
                    default_source_year=default_source_year,
                    custom_factor_kgco2e_per_unit=custom_factor,
                    custom_source_name=custom_source_name,
                    custom_source_url=custom_source_url,
                    custom_source_year=custom_source_year,
                )
            )

        return rows
    finally:
        workbook.close()


def find_sheet_row(
    rows: list[FactorSheetRow],
    *,
    category: str,
    unit_activity: str,
    region: str,
    require_custom_override: bool = False,
) -> FactorSheetRow | None:
    target_category = category.strip()
    target_unit = unit_activity.strip()
    target_region = (region or "").strip().upper()
    candidates = [target_region] if target_region else []
    if "WORLD" not in candidates:
        candidates.append("WORLD")

    for candidate_region in candidates:
        for row in rows:
            if row.category != target_category:
                continue
            if row.unit_activity != target_unit:
                continue
            if row.region != candidate_region:
                continue
            if require_custom_override and not row.has_custom_override:
                continue
            return row

    return None
